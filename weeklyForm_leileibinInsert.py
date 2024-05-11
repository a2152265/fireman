import os
import shutil
import datetime
import openpyxl
import re
from openpyxl.styles import PatternFill, Font

# 1. 找到名稱中包含"訓練問卷調查表"的xlsx檔案，並取得 className
source_files = [file for file in os.listdir() if file.endswith('.xlsx') and '訓練問卷調查表' in file]
if not source_files:
    print("找不到符合條件的xlsx檔案")
    exit()

source_file = source_files[0]
# className = source_file.split('(第2教授班)')[0].split('(回覆')[1]
className = input("請輸入文字：")
print(className)

# 2. 打開檔案並找到"疲勞量表"欄位
workbook = openpyxl.load_workbook(source_file)
sheet = workbook.active
column_student = None
column_fatigue = None
for cell in sheet[1]:
    if "學號" in str(cell.value):
        column_student = cell.column
    if "疲勞量表" in str(cell.value):
        column_fatigue = cell.column
if column_student is None or column_fatigue is None:
    print("找不到'學號'或'疲勞量表'欄位")
    workbook.close()
    exit()

# 3. 取得學號與疲勞量表的陣列，檢查是否為數字
student_numbers = []
fatigue_scores = []
for row in range(2, sheet.max_row + 1):
    student_number = str(sheet.cell(row=row, column=column_student).value).lstrip("0")
    fatigue_score_raw = str(sheet.cell(row=row, column=column_fatigue).value)
    fatigue_score_match = re.search(r'\d+', fatigue_score_raw)
    if fatigue_score_match:
        fatigue_score = fatigue_score_match.group()
    else:
        print("疲勞量表无法提取有效数字:", fatigue_score_raw)
        workbook.close()
        exit()
    if not student_number.isdigit() or not fatigue_score.isdigit():
        print("有人壞壞亂填心情數字或學號", student_number, fatigue_score)
        workbook.close()
        exit()
    student_numbers.append(int(student_number))
    fatigue_scores.append(int(fatigue_score))
    
# 4. 關閉檔案
workbook.close()

# 5. 檢查目錄"疲勞量表"中的xlsx檔案名稱，是否包含47-90(不含55)，共43個檔案，並且在47-90的區間內，如果沒有則噴錯
files_in_folder = os.listdir("疲勞量表")
if len(files_in_folder) != 43:
    print("疲勞量表資料夾中的檔案數量不正確")
    exit()
for number in range(47, 91):
    if number != 55 and f"{number}.xlsx" not in files_in_folder:
        print(f"疲勞量表資料夾中缺少檔案{number}.xlsx")
        exit()

# 6. 複製名稱為"疲勞量表"的資料夾，於"疲勞量表_備份"資料夾下貼上，並命名為"疲勞量表_當天日期YYYYMMDD"
# 修改部分：
backup_folder = "疲勞量表_備份"
if not os.path.exists(backup_folder):
    os.makedirs(backup_folder)
backup_folder_name = datetime.datetime.now().strftime("%Y%m%d")
new_backup_folder_name = backup_folder_name
n = 2
while os.path.exists(os.path.join(backup_folder, new_backup_folder_name)):
    new_backup_folder_name = f"{backup_folder_name}_{n}"
    n += 1
shutil.copytree("疲勞量表", os.path.join(backup_folder, new_backup_folder_name), dirs_exist_ok=True)


# 7. 檢查備份資料夾下層目錄下，包含YYYYMMDD名稱的資料夾數目超過15個的話，則從最舊的開始刪除
subfolders = [f for f in os.listdir(backup_folder) if os.path.isdir(os.path.join(backup_folder, f))]
if len(subfolders) > 9:
    subfolders.sort()
    shutil.rmtree(os.path.join(backup_folder, subfolders[0]))

# 8. 回到"疲勞量表資料夾"，依照陣列的學號，分別開啟對應名稱的xlsx，並都進行以下動作
for student_number, fatigue_score in zip(student_numbers, fatigue_scores):
    filename = f"{student_number}.xlsx"
    filepath = os.path.join("疲勞量表", filename)
    if not os.path.exists(filepath):
        print(f"找不到檔案{filename}")
        exit()
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook.active

    # a. 在A欄位的內容中，依次往下找到內容中包含最新日期的列，並準備在該列底下新增一列
    latest_row = None
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value is not None:
            latest_row = row
    new_row = latest_row + 1 if latest_row is not None else 1

    # b. 新增的列中，分別新增A欄，B-K欄，L欄
    start_date = datetime.datetime.now() - datetime.timedelta(days=datetime.datetime.now().weekday())
    end_date = start_date + datetime.timedelta(days=6)
    week_range = f"{start_date.strftime('%m/%d')}-{end_date.strftime('%m/%d')} {className}"
    sheet.cell(row=new_row, column=1, value=week_range)


    sheet.cell(row=new_row, column=1).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for col in range(2, 12):
# 修改部分：
        if fatigue_score == 1:
            sheet.cell(row=new_row, column=2).fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # B欄
        elif fatigue_score == 2:
            sheet.cell(row=new_row, column=2).fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # B欄
            sheet.cell(row=new_row, column=3).fill = PatternFill(start_color="66FF66", end_color="66FF66", fill_type="solid")  # C欄
        elif 3 <= fatigue_score <= 11:
            fill_colors = [
        "00FF00",  # 綠色
        "66FF66",  # 淺綠色
        "00CC00",  # 深綠色
        "00FFFF",  # 青色
        "66CCFF",  # 淺藍色
        "0000FF",  # 藍色
        "0000CC",  # 深藍色
        "FF9999",  # 淺紅色
        "FF0000",  # 紅色
        "CC0000"   # 深紅色
            ]
            for i in range(2, fatigue_score + 2):
                sheet.cell(row=new_row, column=i).fill = PatternFill(start_color=fill_colors[i - 2], end_color=fill_colors[i - 2], fill_type="solid")

    # L欄內容依據疲勞量表分數，並將底色調整至與L6欄位相同
    if fatigue_score == 1 or fatigue_score == 2:
        l_value = "*太輕鬆囉，週末去動動la~"
    elif fatigue_score == 9 or fatigue_score == 10:
        l_value = "*太累囉，週末多去休息~"
    else:
        l_value = ""
    sheet.cell(row=new_row, column=12).fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    sheet.cell(row=new_row, column=12).value = l_value


    # c. 儲存檔案後離開
    workbook.save(filepath)
    workbook.close()

# 9. 檢查疲勞量表目錄下的檔案內容是否都有在今天更新過，如果沒有則噴錯
for file in os.listdir("疲勞量表"):
    filepath = os.path.join("疲勞量表", file)
    if os.path.getmtime(filepath) < datetime.datetime.now().timestamp() - 86400:
        print(f"檔案{file}未在今天更新過")
        exit()

print("程式執行完成")
