import os
import openpyxl

# 1. 找到名稱中包含"訓練問卷調查表"的xlsx檔案
files = [file for file in os.listdir() if file.endswith('.xlsx') and '訓練問卷調查表' in file]
if not files:
    print("找不到符合條件的xlsx檔案")
    exit()

filename = files[0]

# 2. 打開檔案
workbook = openpyxl.load_workbook(filename)
sheet = workbook.active

# 3. 找到包含"學號"的欄位
student_number_column = None
for cell in sheet[1]:
    if "學號" in str(cell.value):
        student_number_column = cell.column
        break

if student_number_column is None:
    print("找不到包含'學號'的欄位")
    exit()

# 4. 將該欄的所有值存為陣列，並檢查是否為數字
student_numbers = []
for cell in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=student_number_column, max_col=student_number_column, values_only=True):
    if cell[0] is not None:
        if not str(cell[0]).isdigit():
            print("有人壞壞亂填學號", cell[0])
            exit()
        student_numbers.append(int(cell[0]))

# 5. 檢查該陣列是否在47-90之間
for number in student_numbers:
    if not 47 <= number <= 90:
        print("有人壞壞亂填學號", number)
        exit()

# 6. 檢查該陣列的值是否包含47-90所有數字(排除55)
missing_numbers = set(range(47, 91)) - set(student_numbers) - {55}
if missing_numbers:
    for number in missing_numbers:
        print("有人壞壞還沒填", number)
    exit()

print("所有學號填寫正確")
